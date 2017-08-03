import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook
from openpyxl.styles.colors import RED, GREEN, DARKGREEN
from openpyxl.writer.write_only import WriteOnlyCell
import sys
import time

def getXLfiles():
    print("Files Found:")
    files = glob.glob("*.xlsx")
    for i in range(0, len(files)):
        print("[%d] %s" % (i, files[i]))

    UserInputValid = False

    while(not UserInputValid):
        userinput = input("Please Select Old Excel File (Enter Assigned Number):")
        try:
            temp = int(userinput)
            if temp in range(0, len(files)):
                print(files[temp], "was selected\n")
                UserInputValid = True
                oldfile = files[temp]
            else:
                print("Please Enter Number within Range of Files\n")
        except ValueError:
            print("Please Enter a Number\n")

    UserInputValid = False

    while(not UserInputValid):
        userinput = input("Now Select New Excel File (Enter Assigned Number):")
        try:
            temp = int(userinput)
            if temp in range(0, len(files)):
                print(files[temp], "was selected\n")
                UserInputValid = True
                newfile = files[temp]
                time.sleep(1)
            else:
                print("Please Enter Number within Range of Files\n")
        except ValueError:
            print("Please Enter a Number\n")

    if newfile == oldfile:
        print("You Must Select Different Files!")
        print("Program Will Close In 3 Seconds")
        time.sleep(3)
        sys.exit()
    os.system('cls')
    print("Old File: ", oldfile)
    print("New File: ", newfile)
    print("\n")
    return oldfile, newfile



def loadexcel(name):
    data = []
    try:
        wb = load_workbook(name)
    except:
        print("Please rename excel document to data.xlsx")
        print("Program will close in 10 seconds")
        time.sleep(10)
        sys.exit()

    first_sheet = wb.get_sheet_names()[0]
    worksheet = wb.get_sheet_by_name(first_sheet)
    cell_title = worksheet['O1']
    cell_title.value = "Room"
    cell_title.font = Font(bold=True)
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    for row in range(2, worksheet.max_row + 1):
        id = (worksheet.cell(row=row, column=1).value)
        data.append([row, id])
    return data, wb, worksheet

class Excel:
    def __init__(self, fileName):
        self.filename = fileName
        self.wb = None
        self.ws = None
        self.ids = {}
        self.updated = None
        self.removed = None
        self.new = None

    def createheaders(self, sheet, type):
        title = []

        title.append(WriteOnlyCell(sheet, "Changes"))
        title.append(WriteOnlyCell(sheet, "Employee ID"))
        title.append(WriteOnlyCell(sheet, "Last Name"))
        title.append(WriteOnlyCell(sheet, "First Name"))
        title.append(WriteOnlyCell(sheet, "PL Name"))
        title.append(WriteOnlyCell(sheet, "Sex"))
        title.append(WriteOnlyCell(sheet, "Birthdate"))
        title.append(WriteOnlyCell(sheet, "Process Level"))
        title.append(WriteOnlyCell(sheet, "Department"))
        title.append(WriteOnlyCell(sheet, "R Name"))
        title.append(WriteOnlyCell(sheet, "Date Hired"))

        for i in title:
            i.font = Font(bold=True)

        if type == "Updated":
            sheet.append(title)
        else:
            sheet.append(title[1:])

    def setupColumnSize(self):
        self.updated.column_dimensions["A"].width = 20
        self.updated.column_dimensions["B"].width = 15
        self.updated.column_dimensions["C"].width = 30
        self.updated.column_dimensions["D"].width = 20
        self.updated.column_dimensions["E"].width = 35
        self.updated.column_dimensions["F"].width = 5
        self.updated.column_dimensions["G"].width = 20
        self.updated.column_dimensions["H"].width = 15
        self.updated.column_dimensions["I"].width = 15
        self.updated.column_dimensions["J"].width = 40
        self.updated.column_dimensions["K"].width = 20

        self.removed.column_dimensions["A"].width = 15
        self.removed.column_dimensions["B"].width = 30
        self.removed.column_dimensions["C"].width = 20
        self.removed.column_dimensions["D"].width = 35
        self.removed.column_dimensions["E"].width = 5
        self.removed.column_dimensions["F"].width = 20
        self.removed.column_dimensions["G"].width = 15
        self.removed.column_dimensions["H"].width = 15
        self.removed.column_dimensions["I"].width = 40
        self.removed.column_dimensions["J"].width = 20

        self.new.column_dimensions["A"].width = 15
        self.new.column_dimensions["B"].width = 30
        self.new.column_dimensions["C"].width = 20
        self.new.column_dimensions["D"].width = 35
        self.new.column_dimensions["E"].width = 5
        self.new.column_dimensions["F"].width = 20
        self.new.column_dimensions["G"].width = 15
        self.new.column_dimensions["H"].width = 15
        self.new.column_dimensions["I"].width = 40
        self.new.column_dimensions["J"].width = 20

    def setupfile(self):
        self.wb = Workbook(write_only=True)
        self.updated = self.wb.create_sheet("Updated")
        self.removed = self.wb.create_sheet("Removed")
        self.new = self.wb.create_sheet("New")

        self.setupColumnSize()

        self.createheaders(self.updated, "Updated")
        self.createheaders(self.removed, "Removed")
        self.createheaders(self.new, "New")


    def add_new(self, data):
        self.new.append(data)

    def add_removed(self, data):
        self.removed.append(data)

    def add_updated(self, dataold, datanew, change):

        for i in change:
            temp_old = WriteOnlyCell(self.updated, dataold[i])
            temp_new = WriteOnlyCell(self.updated, datanew[i])

            temp_old.font = Font(color=RED)
            temp_new.font = Font(color="008000")

            dataold[i] = temp_old
            datanew[i] = temp_new

        self.updated.append(dataold)
        self.updated.append(datanew)
        self.updated.append([""])

    def loadfile(self):
        self.wb = load_workbook(self.filename, read_only=True)
        first_sheet = self.wb.get_sheet_names()[0]
        self.ws = self.wb.get_sheet_by_name(first_sheet)

    def getRow(self, rownumber):
        data = []
        for column in range(1, self.ws.max_column + 1):
            temp = self.ws.cell(row=rownumber, column=column).value
            data.append(temp)
        return data

    def getIDs(self):
        rownum = 1
        for row in self.ws.rows:
            # Skip headers
            if rownum == 1:
                rownum += 1
                continue
            # Get ID Value
            id = row[0].value

            # Reached end of data if ID is none
            if id is None:
                break

            # Collect Data from row
            rowdata = []
            for cell in row:
                if cell.value == None:
                    break;
                rowdata.append(cell.value)

            self.ids[id] = rowdata
            rownum += 1

    def savefile(self):
        self.wb.save(self.filename + ".xlsx")